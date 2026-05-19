from calendar import monthrange
from datetime import date, timedelta
from decimal import Decimal
import io
from xml.sax.saxutils import escape

from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.budgets.models import Budget
from apps.transactions.models import Transaction


def month_start_for(value):
    return value.replace(day=1)


def next_month_for(value):
    if value.month == 12:
        return value.replace(year=value.year + 1, month=1, day=1)
    return value.replace(month=value.month + 1, day=1)


def add_months(value, months):
    month_index = value.year * 12 + value.month - 1 + months
    return date(month_index // 12, month_index % 12 + 1, 1)


def quarter_start_for(value):
    return value.replace(month=((value.month - 1) // 3) * 3 + 1, day=1)


def year_start_for(value):
    return value.replace(month=1, day=1)


def parse_month_param(raw_value):
    if not raw_value:
        return month_start_for(timezone.localdate())
    if len(raw_value) == 7:
        year, month = raw_value.split("-")
        return date(int(year), int(month), 1)
    return month_start_for(date.fromisoformat(raw_value))


def parse_date_param(raw_value):
    return date.fromisoformat(raw_value) if raw_value else None


def parse_reference_date(raw_value):
    if not raw_value:
        return timezone.localdate()
    if len(raw_value) == 7:
        return parse_month_param(raw_value)
    return date.fromisoformat(raw_value)


def format_report_date(value):
    return value.strftime("%d-%m-%Y")


def format_money(value):
    numeric = value or Decimal("0")
    return round(float(numeric), 2)


def format_money_display(value):
    numeric = Decimal(value or "0")
    return f"Rs. {numeric:,.2f}"


def period_label(period):
    return {
        "daily": "Daily",
        "weekly": "Weekly",
        "monthly": "Monthly",
        "quarterly": "Quarterly",
        "yearly": "Yearly",
        "custom": "Custom",
    }.get(period, "Transactions")


def resolve_export_window(params):
    today = timezone.localdate()
    month = params.get("month")
    start_date = parse_date_param(params.get("start_date"))
    end_date = parse_date_param(params.get("end_date"))
    period = (params.get("period") or "").strip().lower()
    reference_date = parse_reference_date(params.get("reference_date") or params.get("month"))

    if start_date or end_date:
        return start_date, end_date or today, period or "custom"
    if month:
        month_start = parse_month_param(month)
        month_end = min(next_month_for(month_start) - timedelta(days=1), today)
        return month_start, month_end, "monthly"
    if period in {"daily", "weekly", "monthly", "quarterly", "yearly"}:
        period_start, period_end = resolve_period_range(period, reference_date)
        return period_start, period_end, period
    return None, today, "custom"


def build_report_title(period):
    label = period_label(period)
    return f"{label} Transactions Report" if label != "Transactions" else "Transactions Report"


def build_report_filename(period, start_date, end_date):
    period_slug = period or "custom"
    if start_date and end_date:
        return f"transactions-{period_slug}-{start_date.isoformat()}-to-{end_date.isoformat()}"
    if end_date:
        return f"transactions-{period_slug}-through-{end_date.isoformat()}"
    return f"transactions-{period_slug}-report"


def build_filters_summary(period, start_date, end_date, tx_type):
    type_label = {"income": "Income only", "expense": "Expense only"}.get(tx_type, "All transactions")
    period_text = period_label(period)
    if start_date and end_date:
        return f"{period_text} range: {format_report_date(start_date)} to {format_report_date(end_date)} | {type_label}"
    if end_date:
        return f"{period_text} range: Beginning to {format_report_date(end_date)} | {type_label}"
    return f"{period_text} transactions | {type_label}"


def resolve_period_range(period, reference):
    today = timezone.localdate()

    if period == "daily":
        return reference, reference
    if period == "weekly":
        start_date = reference - timedelta(days=reference.weekday())
        return start_date, min(start_date + timedelta(days=6), today)
    if period == "monthly":
        start_date = month_start_for(reference)
        return start_date, min(next_month_for(start_date) - timedelta(days=1), today)
    if period == "quarterly":
        start_date = quarter_start_for(reference)
        return start_date, min(add_months(start_date, 3) - timedelta(days=1), today)
    if period == "yearly":
        start_date = year_start_for(reference)
        return start_date, min(date(reference.year, 12, 31), today)
    return None, None


def apply_transaction_filters(queryset, params):
    tx_type = params.get("type")
    category_id = params.get("category")
    start_date = parse_date_param(params.get("start_date"))
    end_date = parse_date_param(params.get("end_date"))
    search = params.get("search")
    month = params.get("month")
    period = (params.get("period") or "").strip().lower()
    reference_date = parse_reference_date(params.get("reference_date") or params.get("month"))

    if tx_type in {"income", "expense"}:
        queryset = queryset.filter(type=tx_type)
    if category_id:
        queryset = queryset.filter(category_id=category_id)
    if period in {"daily", "weekly", "monthly", "quarterly", "yearly"} and not start_date and not end_date:
        start_date, end_date = resolve_period_range(period, reference_date)
    elif month:
        month_start = parse_month_param(month)
        queryset = queryset.filter(date__gte=month_start, date__lt=next_month_for(month_start))
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    if search:
        queryset = queryset.filter(
            Q(title__icontains=search)
            | Q(note__icontains=search)
            | Q(category__name__icontains=search)
        )
    return queryset


def serialize_transaction(transaction):
    return {
        "id": transaction.id,
        "type": transaction.type,
        "amount": str(transaction.amount),
        "category": transaction.category_id,
        "category_name": transaction.category.name if transaction.category else "Uncategorized",
        "title": transaction.title,
        "note": transaction.note,
        "date": transaction.date.isoformat(),
        "display_date": format_report_date(transaction.date),
        "created_at": transaction.created_at.isoformat(),
        "updated_at": transaction.updated_at.isoformat(),
    }


def transaction_delta(transaction):
    return transaction.amount if transaction.type == "income" else -transaction.amount


def opening_balance_before(user, transaction):
    earlier_transactions = Transaction.objects.filter(user=user).filter(
        Q(date__lt=transaction.date)
        | Q(date=transaction.date, created_at__lt=transaction.created_at)
        | Q(date=transaction.date, created_at=transaction.created_at, id__lt=transaction.id)
    )
    return aggregate_total(earlier_transactions, "income") - aggregate_total(earlier_transactions, "expense")


def build_transaction_report_rows(user, transactions):
    if not transactions:
        return []

    running_balance = opening_balance_before(user, transactions[0])
    rows = []

    for transaction in transactions:
        running_balance += transaction_delta(transaction)
        row = serialize_transaction(transaction)
        row["available_balance"] = str(running_balance)
        rows.append(row)

    return rows


def aggregate_total(queryset, tx_type):
    return queryset.filter(type=tx_type).aggregate(total=Sum("amount")).get("total") or Decimal("0")


def budget_snapshot_for(user, month):
    budget = Budget.objects.filter(user=user, month=month).first()
    if not budget:
        return None

    spent = aggregate_total(
        Transaction.objects.filter(
            user=user,
            type="expense",
            date__gte=month,
            date__lt=next_month_for(month),
        ),
        "expense",
    )
    remaining = budget.limit_amount - spent
    progress = round(float((spent / budget.limit_amount) * 100), 2) if budget.limit_amount else 0
    if spent > budget.limit_amount:
        status = "exceeded"
    elif budget.limit_amount and spent / budget.limit_amount >= Decimal("0.8"):
        status = "near_limit"
    else:
        status = "healthy"

    return {
        "id": budget.id,
        "month": budget.month.isoformat(),
        "limit_amount": str(budget.limit_amount),
        "spent_amount": str(spent),
        "remaining_amount": str(remaining),
        "progress_percent": progress,
        "status": status,
    }


def build_dashboard_payload(user):
    today = timezone.localdate()
    month = month_start_for(today)
    all_transactions = Transaction.objects.filter(user=user).select_related("category")
    current_month_transactions = all_transactions.filter(date__gte=month, date__lt=next_month_for(month))

    total_income = aggregate_total(all_transactions, "income")
    total_expense = aggregate_total(all_transactions, "expense")
    month_income = aggregate_total(current_month_transactions, "income")
    month_expense = aggregate_total(current_month_transactions, "expense")

    category_rows = (
        current_month_transactions.filter(type="expense")
        .values("category__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )
    category_total = sum((row["total"] for row in category_rows), Decimal("0"))
    category_breakdown = [
        {
            "name": row["category__name"] or "Uncategorized",
            "amount": format_money(row["total"]),
            "percentage": round(float((row["total"] / category_total) * 100), 2) if category_total else 0,
        }
        for row in category_rows
    ]

    trend = []
    current_index = month.year * 12 + month.month - 1
    for offset in range(5, -1, -1):
        month_index = current_index - offset
        trend_year = month_index // 12
        trend_month = month_index % 12 + 1
        trend_start = date(trend_year, trend_month, 1)
        trend_queryset = all_transactions.filter(date__gte=trend_start, date__lt=next_month_for(trend_start))
        trend.append(
            {
                "month": trend_start.strftime("%b %Y"),
                "income": format_money(aggregate_total(trend_queryset, "income")),
                "expense": format_money(aggregate_total(trend_queryset, "expense")),
            }
        )

    return {
        "totals": {
            "income": format_money(total_income),
            "expense": format_money(total_expense),
            "balance": format_money(total_income - total_expense),
        },
        "current_month": {
            "month": month.strftime("%B %Y"),
            "income": format_money(month_income),
            "expense": format_money(month_expense),
            "balance": format_money(month_income - month_expense),
            "savings_rate": round(float(((month_income - month_expense) / month_income) * 100), 2)
            if month_income
            else 0,
        },
        "recent_transactions": [
            serialize_transaction(transaction)
            for transaction in all_transactions.order_by("-date", "-created_at")[:5]
        ],
        "category_breakdown": category_breakdown,
        "monthly_trend": trend,
        "budget": budget_snapshot_for(user, month),
    }


def build_monthly_report_payload(user, month):
    next_month = next_month_for(month)
    transactions = (
        Transaction.objects.filter(user=user, date__gte=month, date__lt=next_month)
        .select_related("category")
        .order_by("date", "created_at", "id")
    )
    income_total = aggregate_total(transactions, "income")
    expense_total = aggregate_total(transactions, "expense")

    return {
        "month": month.strftime("%B %Y"),
        "month_start": month.isoformat(),
        "month_end": date(month.year, month.month, monthrange(month.year, month.month)[1]).isoformat(),
        "totals": {
            "income": format_money(income_total),
            "expense": format_money(expense_total),
            "balance": format_money(income_total - expense_total),
        },
        "budget": budget_snapshot_for(user, month),
        "transactions": build_transaction_report_rows(user, list(transactions)),
    }


def build_category_summary_payload(user, month):
    transactions = Transaction.objects.filter(
        user=user,
        type="expense",
        date__gte=month,
        date__lt=next_month_for(month),
    )
    rows = transactions.values("category__name").annotate(total=Sum("amount")).order_by("-total")
    grand_total = sum((row["total"] for row in rows), Decimal("0"))
    categories = [
        {
            "name": row["category__name"] or "Uncategorized",
            "amount": format_money(row["total"]),
            "percentage": round(float((row["total"] / grand_total) * 100), 2) if grand_total else 0,
        }
        for row in rows
    ]
    return {
        "month": month.strftime("%B %Y"),
        "total_expense": format_money(grand_total),
        "categories": categories,
    }


def build_csv_response_rows(transactions, totals=None):
    totals = totals or {
        "income": sum(float(transaction["amount"]) for transaction in transactions if transaction["type"] == "income"),
        "expense": sum(float(transaction["amount"]) for transaction in transactions if transaction["type"] == "expense"),
    }

    rows = [["Date", "Type", "Category", "Title", "Note", "Amount", "Available Balance"]]
    rows.extend(
        [
            transaction["display_date"],
            transaction["type"].title(),
            transaction["category_name"],
            transaction["title"],
            transaction["note"],
            transaction["amount"],
            transaction["available_balance"],
        ]
        for transaction in transactions
    )

    rows.append([])
    rows.append(["Visual Summary - Totals"])
    rows.append(["Metric", "Amount"])
    rows.extend(
        [
            ["Income", f"{float(totals['income']):.2f}"],
            ["Expense", f"{float(totals['expense']):.2f}"],
            ["Net Balance", f"{float(totals['income'] - totals['expense']):.2f}"],
        ]
    )

    rows.append([])
    rows.append(["Visual Summary - Expense Categories"])
    rows.append(["Category", "Amount"])
    category_rows = build_expense_category_summary(transactions, limit=None)
    if category_rows:
        rows.extend([[name, f"{amount:.2f}"] for name, amount in category_rows])
    else:
        rows.append(["No expense categories", "0.00"])

    rows.append([])
    rows.append(["Visual Summary - Daily Cash Flow"])
    rows.append(["Date", "Income", "Expense", "Net"])
    daily_rows = build_daily_cashflow_summary(transactions)
    if daily_rows:
        rows.extend(
            [
                [display_date, f"{income:.2f}", f"{expense:.2f}", f"{net:.2f}"]
                for display_date, income, expense, net in daily_rows
            ]
        )
    else:
        rows.append(["No transaction days", "0.00", "0.00", "0.00"])

    return rows


def build_expense_category_summary(transactions, limit=6):
    category_totals = {}
    for transaction in transactions:
        if transaction["type"] != "expense":
            continue
        category_name = transaction["category_name"] or "Uncategorized"
        category_totals[category_name] = category_totals.get(category_name, 0.0) + float(transaction["amount"])

    sorted_categories = sorted(category_totals.items(), key=lambda item: item[1], reverse=True)
    if limit and len(sorted_categories) > limit:
        visible = sorted_categories[: limit - 1]
        other_total = sum(amount for _, amount in sorted_categories[limit - 1 :])
        sorted_categories = visible + [("Other", other_total)]
    return sorted_categories


def build_daily_cashflow_summary(transactions):
    daily_totals = {}
    for transaction in transactions:
        key = transaction["date"]
        if key not in daily_totals:
            daily_totals[key] = {
                "display_date": transaction["display_date"],
                "income": 0.0,
                "expense": 0.0,
            }

        amount = float(transaction["amount"])
        if transaction["type"] == "income":
            daily_totals[key]["income"] += amount
        else:
            daily_totals[key]["expense"] += amount

    rows = []
    for key in sorted(daily_totals):
        item = daily_totals[key]
        rows.append(
            (
                item["display_date"],
                item["income"],
                item["expense"],
                item["income"] - item["expense"],
            )
        )
    return rows


def build_report_visual_snapshot(transactions, totals):
    category_summary = build_expense_category_summary(transactions)
    return {
        "totals_chart": [
            ("Income", float(totals["income"])),
            ("Expense", float(totals["expense"])),
            ("Net Balance", float(totals["income"] - totals["expense"])),
        ],
        "category_chart": category_summary,
    }


def append_workbook_visuals(workbook, visual_snapshot):
    visuals_sheet = workbook.create_sheet("Visuals")
    visuals_sheet.sheet_view.showGridLines = False

    visuals_sheet["A1"] = "Visual Summary"
    visuals_sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="231A12")
    visuals_sheet["A3"] = "Totals Snapshot"
    visuals_sheet["A3"].font = Font(name="Calibri", size=12, bold=True, color="231A12")
    visuals_sheet.append([])
    visuals_sheet.append(["Metric", "Amount"])

    totals_start_row = 5
    for label, amount in visual_snapshot["totals_chart"]:
        visuals_sheet.append([label, amount])

    for row in range(totals_start_row, totals_start_row + len(visual_snapshot["totals_chart"])):
        visuals_sheet[f"B{row}"].number_format = '#,##0.00'

    totals_chart = BarChart()
    totals_chart.type = "col"
    totals_chart.style = 10
    totals_chart.title = "Income vs Expense vs Net"
    totals_chart.y_axis.title = "Amount"
    totals_chart.height = 7
    totals_chart.width = 12
    totals_chart.add_data(Reference(visuals_sheet, min_col=2, min_row=4, max_row=4 + len(visual_snapshot["totals_chart"])), titles_from_data=True)
    totals_chart.set_categories(Reference(visuals_sheet, min_col=1, min_row=5, max_row=4 + len(visual_snapshot["totals_chart"])))
    visuals_sheet.add_chart(totals_chart, "D3")

    category_anchor_row = 11
    visuals_sheet[f"A{category_anchor_row}"] = "Expense Categories"
    visuals_sheet[f"A{category_anchor_row}"].font = Font(name="Calibri", size=12, bold=True, color="231A12")
    visuals_sheet.append([])
    visuals_sheet.append(["Category", "Amount"])

    category_table_start = category_anchor_row + 2
    for label, amount in visual_snapshot["category_chart"]:
        visuals_sheet.append([label, amount])

    category_end_row = category_table_start + len(visual_snapshot["category_chart"]) - 1
    for row in range(category_table_start, category_end_row + 1):
        visuals_sheet[f"B{row}"].number_format = '#,##0.00'

    if visual_snapshot["category_chart"]:
        pie_chart = PieChart()
        pie_chart.title = "Expense Mix"
        pie_chart.height = 8
        pie_chart.width = 11
        pie_chart.add_data(Reference(visuals_sheet, min_col=2, min_row=category_anchor_row + 1, max_row=category_end_row), titles_from_data=True)
        pie_chart.set_categories(Reference(visuals_sheet, min_col=1, min_row=category_table_start, max_row=category_end_row))
        visuals_sheet.add_chart(pie_chart, "D14")

    visuals_sheet.column_dimensions["A"].width = 22
    visuals_sheet.column_dimensions["B"].width = 16
    return visuals_sheet


def build_pdf_visual_summary(visual_snapshot):
    drawing = Drawing(730, 190)
    drawing.add(String(0, 172, "Visual Snapshot", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor("#231A12")))

    max_total_value = max((amount for _, amount in visual_snapshot["totals_chart"]), default=1.0) or 1.0
    drawing.add(String(0, 150, "Totals", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#5A4633")))
    total_colors = ["#B6D5BF", "#E9CC79", "#D8897A"]
    bar_origin_x = 0
    bar_origin_y = 132
    bar_width = 220
    bar_height = 14
    for index, (label, amount) in enumerate(visual_snapshot["totals_chart"]):
        y = bar_origin_y - index * 34
        drawing.add(String(bar_origin_x, y + 18, label, fontName="Helvetica", fontSize=9, fillColor=colors.HexColor("#231A12")))
        drawing.add(Rect(bar_origin_x, y, bar_width, bar_height, fillColor=colors.HexColor("#F6EFE4"), strokeColor=colors.HexColor("#D7C9B1"), strokeWidth=0.5))
        fill_width = max(6, (abs(amount) / max_total_value) * bar_width) if amount else 0
        drawing.add(Rect(bar_origin_x, y, fill_width, bar_height, fillColor=colors.HexColor(total_colors[index % len(total_colors)]), strokeColor=None))
        drawing.add(String(bar_origin_x + bar_width + 10, y + 3, format_money_display(amount), fontName="Helvetica", fontSize=9, fillColor=colors.HexColor("#231A12")))

    drawing.add(String(360, 150, "Top Expense Categories", fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#5A4633")))
    max_category_value = max((amount for _, amount in visual_snapshot["category_chart"]), default=1.0) or 1.0
    category_origin_x = 360
    category_origin_y = 132
    category_width = 200
    for index, (label, amount) in enumerate(visual_snapshot["category_chart"][:5]):
        y = category_origin_y - index * 24
        drawing.add(String(category_origin_x, y + 2, label[:24], fontName="Helvetica", fontSize=8.5, fillColor=colors.HexColor("#231A12")))
        drawing.add(Rect(category_origin_x + 92, y, category_width, 10, fillColor=colors.HexColor("#F6EFE4"), strokeColor=None))
        fill_width = max(4, (amount / max_category_value) * category_width) if amount else 0
        drawing.add(Rect(category_origin_x + 92, y, fill_width, 10, fillColor=colors.HexColor("#B6D5BF"), strokeColor=None))
        drawing.add(String(category_origin_x + 300, y + 1, format_money_display(amount), fontName="Helvetica", fontSize=8.5, fillColor=colors.HexColor("#5A4633")))

    if not visual_snapshot["category_chart"]:
        drawing.add(String(360, 124, "No expense category data for this export.", fontName="Helvetica", fontSize=8.5, fillColor=colors.HexColor("#5A4633")))

    return drawing


def build_workbook(transactions, title, filters_summary, totals):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    accent_fill = PatternFill("solid", fgColor="D9EAD3")
    soft_fill = PatternFill("solid", fgColor="F6EFE4")
    header_fill = PatternFill("solid", fgColor="1F1B16")
    border = Border(
        left=Side(style="thin", color="D7C9B1"),
        right=Side(style="thin", color="D7C9B1"),
        top=Side(style="thin", color="D7C9B1"),
        bottom=Side(style="thin", color="D7C9B1"),
    )

    sheet.merge_cells("A1:G1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="231A12")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A2:G2")
    sheet["A2"] = filters_summary
    sheet["A2"].font = Font(name="Calibri", size=11, italic=True, color="5A4633")
    sheet["A2"].alignment = Alignment(horizontal="center")

    sheet["A4"] = "Income"
    sheet["B4"] = "Expense"
    sheet["C4"] = "Net Balance"
    sheet["A5"] = float(totals["income"])
    sheet["B5"] = float(totals["expense"])
    sheet["C5"] = float(totals["income"] - totals["expense"])

    for cell_ref in ("A4", "B4", "C4"):
        sheet[cell_ref].font = Font(bold=True, color="231A12")
        sheet[cell_ref].fill = accent_fill
        sheet[cell_ref].alignment = Alignment(horizontal="center")
        sheet[cell_ref].border = border

    for cell_ref in ("A5", "B5", "C5"):
        sheet[cell_ref].number_format = '#,##0.00'
        sheet[cell_ref].alignment = Alignment(horizontal="center")
        sheet[cell_ref].border = border

    header_row_index = 7
    sheet.append([])
    sheet.append(["Date", "Type", "Category", "Title", "Note", "Amount", "Available Balance"])

    for transaction in transactions:
        sheet.append(
            [
                transaction["display_date"],
                transaction["type"].title(),
                transaction["category_name"],
                transaction["title"],
                transaction["note"],
                float(transaction["amount"]),
                float(transaction["available_balance"]),
            ]
        )

    for cell in sheet[header_row_index]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        row_fill = soft_fill if (row_index - header_row_index) % 2 == 0 else None
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill

        sheet[f"F{row_index}"].number_format = '#,##0.00'
        sheet[f"G{row_index}"].number_format = '#,##0.00'
        sheet[f"A{row_index}"].alignment = Alignment(horizontal="center")
        sheet[f"B{row_index}"].alignment = Alignment(horizontal="center")
        sheet[f"F{row_index}"].alignment = Alignment(horizontal="right")
        sheet[f"G{row_index}"].alignment = Alignment(horizontal="right")

    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:G{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 26
    sheet.column_dimensions["E"].width = 32
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 20
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 20

    append_workbook_visuals(workbook, build_report_visual_snapshot(transactions, totals))

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_pdf_bytes(user, transactions, title, filters_summary, totals):
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=28,
        rightMargin=28,
        topMargin=28,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = []

    title_style = styles["Title"]
    title_style.fontSize = 18
    title_style.leading = 22
    title_style.textColor = colors.HexColor("#231A12")

    meta_style = styles["BodyText"]
    meta_style.fontSize = 9
    meta_style.leading = 12
    meta_style.textColor = colors.HexColor("#5A4633")

    story.append(Paragraph(escape(title), title_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"User: {escape(user.email)}", meta_style))
    story.append(Paragraph(escape(filters_summary), meta_style))
    story.append(Paragraph(f"Generated on {timezone.localtime().strftime('%d-%m-%Y %I:%M %p')}", meta_style))
    story.append(Spacer(1, 12))

    totals_table = Table(
        [
            ["Income", "Expense", "Net Balance"],
            [
                format_money_display(totals["income"]),
                format_money_display(totals["expense"]),
                format_money_display(totals["income"] - totals["expense"]),
            ],
        ],
        colWidths=[150, 150, 170],
    )
    totals_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAD3")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#231A12")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#D7C9B1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7C9B1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(totals_table)
    story.append(Spacer(1, 12))
    story.append(build_pdf_visual_summary(build_report_visual_snapshot(transactions, totals)))
    story.append(Spacer(1, 12))

    table_rows = [[
        "Date",
        "Type",
        "Category",
        "Title",
        "Note",
        "Amount",
        "Available Balance",
    ]]

    for transaction in transactions:
        table_rows.append(
            [
                transaction["display_date"],
                transaction["type"].title(),
                Paragraph(escape(transaction["category_name"]), meta_style),
                Paragraph(escape(transaction["title"]), meta_style),
                Paragraph(escape(transaction["note"] or "-"), meta_style),
                format_money_display(transaction["amount"]),
                format_money_display(transaction["available_balance"]),
            ]
        )

    if len(table_rows) == 1:
        table_rows.append(["-", "-", "-", "No transactions found for the selected range.", "-", "-", "-"])

    transactions_table = Table(
        table_rows,
        colWidths=[68, 58, 118, 128, 165, 88, 98],
        repeatRows=1,
    )
    transactions_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F1B16")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7C9B1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (5, 1), (6, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6EFE4")]),
            ]
        )
    )
    story.append(transactions_table)

    doc.build(story)
    output.seek(0)
    return output
