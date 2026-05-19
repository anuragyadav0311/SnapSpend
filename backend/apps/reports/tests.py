import csv
from datetime import date, timedelta
import io

from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from apps.budgets.models import Budget
from apps.transactions.models import Category, Transaction


User = get_user_model()


class ReportsApiTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(name="Reports User", email="reports@example.com", password="SecurePass@123")
        self.expense_category = Category.objects.create(name="Food & Dining", type="expense", user=None)
        self.income_category = Category.objects.create(name="Salary", type="income", user=None)
        Transaction.objects.create(
            user=self.user,
            type="income",
            amount="90000.00",
            category=self.income_category,
            title="Salary",
            date=date(2026, 5, 1),
        )
        Transaction.objects.create(
            user=self.user,
            type="expense",
            amount="1200.00",
            category=self.expense_category,
            title="Groceries",
            note="Week one",
            date=date(2026, 5, 5),
        )
        Transaction.objects.create(
            user=self.user,
            type="expense",
            amount="800.00",
            category=self.expense_category,
            title="Dining out",
            date=date(2026, 4, 28),
        )
        Budget.objects.create(user=self.user, month=date(2026, 5, 1), limit_amount="5000.00")
        login_response = self.client.post(
            "/api/auth/login/",
            {"email": "reports@example.com", "password": "SecurePass@123"},
            format="json",
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {login_response.data['access']}")

    def test_dashboard_returns_expected_sections(self):
        response = self.client.get("/api/reports/dashboard/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("totals", response.data)
        self.assertIn("recent_transactions", response.data)
        self.assertIn("budget", response.data)

    def test_monthly_report_respects_month_filter(self):
        response = self.client.get("/api/reports/monthly/?month=2026-05")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["month"], "May 2026")
        self.assertEqual(len(response.data["transactions"]), 2)
        self.assertEqual(response.data["transactions"][0]["date"], "2026-05-01")
        self.assertEqual(response.data["transactions"][0]["display_date"], "01-05-2026")
        self.assertEqual(response.data["transactions"][0]["available_balance"], "89200.00")
        self.assertEqual(response.data["transactions"][1]["available_balance"], "88000.00")

    def test_category_summary_returns_sorted_expense_categories(self):
        response = self.client.get("/api/reports/category-summary/?month=2026-05")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["categories"][0]["name"], "Food & Dining")

    def test_export_endpoints_return_expected_content_types(self):
        today = timezone.localdate()
        csv_response = self.client.get("/api/reports/export/csv/?month=2026-05")
        xlsx_response = self.client.get("/api/reports/export/xlsx/?month=2026-05")
        pdf_response = self.client.get("/api/reports/export/pdf/?month=2026-05")

        self.assertEqual(csv_response.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", csv_response["Content-Type"])
        self.assertEqual(xlsx_response.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", xlsx_response["Content-Type"])
        self.assertEqual(pdf_response.status_code, status.HTTP_200_OK)
        self.assertIn("application/pdf", pdf_response["Content-Type"])
        expected_end = min(date(2026, 5, 31), today)
        expected_stem = f"transactions-monthly-2026-05-01-to-{expected_end.isoformat()}"
        self.assertIn(f"{expected_stem}.csv", csv_response["Content-Disposition"])
        self.assertIn(f"{expected_stem}.xlsx", xlsx_response["Content-Disposition"])
        self.assertIn(f"{expected_stem}.pdf", pdf_response["Content-Disposition"])

    def test_csv_export_uses_dd_mm_yyyy_order_and_available_balance(self):
        response = self.client.get("/api/reports/export/csv/?month=2026-05")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        rows = list(csv.reader(io.StringIO(response.content.decode("utf-8"))))

        self.assertEqual(rows[0], ["Date", "Type", "Category", "Title", "Note", "Amount", "Available Balance"])
        self.assertEqual(rows[1], ["01-05-2026", "Income", "Salary", "Salary", "", "90000.00", "89200.00"])
        self.assertEqual(rows[2], ["05-05-2026", "Expense", "Food & Dining", "Groceries", "Week one", "1200.00", "88000.00"])
        self.assertIn(["Visual Summary - Totals"], rows)
        self.assertIn(["Metric", "Amount"], rows)
        self.assertIn(["Income", "90000.00"], rows)
        self.assertIn(["Expense", "1200.00"], rows)
        self.assertIn(["Net Balance", "88800.00"], rows)
        self.assertIn(["Visual Summary - Expense Categories"], rows)
        self.assertIn(["Food & Dining", "1200.00"], rows)
        self.assertIn(["Visual Summary - Daily Cash Flow"], rows)
        self.assertIn(["01-05-2026", "90000.00", "0.00", "90000.00"], rows)
        self.assertIn(["05-05-2026", "0.00", "1200.00", "-1200.00"], rows)

    def test_export_supports_period_reference_date_filters(self):
        response = self.client.get("/api/reports/export/csv/?period=weekly&reference_date=2026-05-05")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        rows = list(csv.reader(io.StringIO(response.content.decode("utf-8"))))

        self.assertEqual(rows[1][0], "05-05-2026")
        self.assertIn(["Visual Summary - Totals"], rows)
        self.assertIn(["Expense", "1200.00"], rows)
        self.assertIn("transactions-weekly-2026-05-04-to-2026-05-10.csv", response["Content-Disposition"])

    def test_excel_export_includes_visuals_sheet(self):
        response = self.client.get("/api/reports/export/xlsx/?month=2026-05")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("Report", workbook.sheetnames)
        self.assertIn("Visuals", workbook.sheetnames)

    def test_report_export_rejects_future_dates(self):
        future_date = timezone.localdate() + timedelta(days=1)
        response = self.client.get(f"/api/reports/export/csv/?start_date={future_date.isoformat()}")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("start_date", response.data)

    def test_monthly_report_rejects_future_month(self):
        future_month = (timezone.localdate() + timedelta(days=31)).replace(day=1)
        response = self.client.get(f"/api/reports/monthly/?month={future_month.isoformat()}")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("month", response.data)
